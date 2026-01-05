from datetime import datetime, timedelta
from typing import Optional, Any, Dict, List, Tuple, Set
from uuid import UUID

from fastapi import APIRouter, Depends, Query, HTTPException
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, text, func, case
import csv
import io
import json

from openpyxl import Workbook

from database.db import get_db
from models.models import FlowLog, Lead, Customer, Message

router = APIRouter(prefix="/api/flow-logs", tags=["flow-logs"])


def _parse_dt(val: Optional[str]) -> Optional[datetime]:
    if not val:
        return None
    try:
        # Accept "YYYY-MM-DD" or ISO strings
        if len(val) == 10:
            # Parse as date and set to start of day (00:00:00) - naive datetime
            dt = datetime.strptime(val, "%Y-%m-%d")
            return dt.replace(hour=0, minute=0, second=0, microsecond=0)
        # For ISO strings, parse and ensure naive datetime
        dt = datetime.fromisoformat(val.replace('Z', '+00:00'))
        # If timezone-aware, convert to naive (assuming UTC, then remove timezone)
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return dt
    except Exception as e:
        print(f"Error parsing date '{val}': {e}")
        return None


def _get_peer_numbers_for_customers(db: Session, wa_ids: List[str]) -> Dict[str, str]:
    """
    Get peer numbers (business WhatsApp numbers) for a list of customer wa_ids.
    Returns a dict mapping customer wa_id to their peer number.
    The peer number is the business number the customer messaged.
    """
    if not wa_ids:
        return {}
    
    # Get the first message for each customer to determine peer number
    # Peer number is the business number (to_wa_id when customer sent, from_wa_id when business sent)
    peer_subq = (
        db.query(
            Message.customer_id,
            func.min(Message.timestamp).label('first_timestamp')
        )
        .join(Customer, Message.customer_id == Customer.id)
        .filter(Customer.wa_id.in_(wa_ids))
        .group_by(Message.customer_id)
        .subquery()
    )
    
    peer_results = (
        db.query(
            Customer.wa_id,
            func.coalesce(
                case(
                    (Message.from_wa_id == Customer.wa_id, Message.to_wa_id),
                    else_=Message.from_wa_id
                ),
                ""
            ).label('peer_number')
        )
        .select_from(Message)
        .join(
            peer_subq,
            and_(
                Message.customer_id == peer_subq.c.customer_id,
                Message.timestamp == peer_subq.c.first_timestamp
            )
        )
        .join(Customer, Message.customer_id == Customer.id)
        .all()
    )
    
    # Return map with all wa_ids, using empty string if peer_number not found
    result = {}
    for wa_id, peer_number in peer_results:
        if wa_id:
            result[wa_id] = peer_number or ""
    
    return result


def _extract_keywords_from_messages(db: Session, wa_ids: List[str]) -> Dict[str, str]:
    """
    Extract keywords from customer messages only (not agent messages).
    Returns a dict mapping customer wa_id to comma-separated keywords found in their messages.
    """
    if not wa_ids:
        return {}
    
    # Define keywords to search for (case-insensitive)
    keywords = [
        "price",
        "appointment",
        "cost",
        "vacancy",
        "contact us",
        "clinic address",
        "clinic number",
        "treatment information",
        "treatment type",
    ]
    
    # Get all messages from customers ONLY (not agent messages)
    messages = (
        db.query(
            Customer.wa_id,
            Message.body
        )
        .join(Message, Customer.id == Message.customer_id)
        .filter(Customer.wa_id.in_(wa_ids))
        .filter(Message.sender_type == "customer")  # ONLY customer messages, not agent
        .filter(Message.body.isnot(None))
        .all()
    )
    
    # Map wa_id to found keywords
    keyword_map: Dict[str, Set[str]] = {}
    for wa_id, body in messages:
        if not wa_id or not body:
            continue
        
        body_lower = body.lower()
        found_keywords = []
        
        for keyword in keywords:
            if keyword.lower() in body_lower:
                # Use the original keyword format for display
                found_keywords.append(keyword)
        
        if found_keywords:
            if wa_id not in keyword_map:
                keyword_map[wa_id] = set()
            keyword_map[wa_id].update(found_keywords)
    
    # Convert sets to comma-separated strings
    return {wa_id: ", ".join(sorted(keywords)) for wa_id, keywords in keyword_map.items()}


def _get_who_ended_chat_first(db: Session, wa_ids: List[str]) -> Dict[str, str]:
    """
    Determine who ended the conversation first (who sent the last message).
    Returns a dict mapping customer wa_id to "Customer" or "Agent".
    """
    if not wa_ids:
        return {}
    
    # Get the last message for each customer (handle multiple messages with same timestamp)
    last_message_subq = (
        db.query(
            Message.customer_id,
            func.max(Message.timestamp).label('last_timestamp'),
            func.max(Message.id).label('last_message_id')  # Use ID as tiebreaker
        )
        .join(Customer, Message.customer_id == Customer.id)
        .filter(Customer.wa_id.in_(wa_ids))
        .group_by(Message.customer_id)
        .subquery()
    )
    
    last_messages = (
        db.query(
            Customer.wa_id,
            Message.sender_type
        )
        .select_from(Message)
        .join(
            last_message_subq,
            and_(
                Message.customer_id == last_message_subq.c.customer_id,
                Message.timestamp == last_message_subq.c.last_timestamp,
                Message.id == last_message_subq.c.last_message_id
            )
        )
        .join(Customer, Message.customer_id == Customer.id)
        .all()
    )
    
    # Map wa_id to who ended the chat
    result = {}
    for wa_id, sender_type in last_messages:
        if not wa_id:
            continue
        if sender_type == "customer":
            result[wa_id] = "Customer"
        elif sender_type:
            result[wa_id] = "Agent"
        else:
            result[wa_id] = "Unknown"
    
    return result


@router.get("")
def list_flow_logs(
    db: Session = Depends(get_db),
    flow_type: Optional[str] = Query(None),
    wa_id: Optional[str] = Query(None),
    step: Optional[str] = Query(None),
    status_code: Optional[int] = Query(None),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
    q: Optional[str] = Query(None, description="search in description"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=5000),
):
    try:
        filters = []
        if flow_type:
            filters.append(FlowLog.flow_type == flow_type)
        if wa_id:
            filters.append(FlowLog.wa_id == wa_id)
        if step:
            filters.append(FlowLog.step == step)
        if status_code is not None:
            filters.append(FlowLog.status_code == status_code)
        if date_from:
            dt_from = _parse_dt(date_from)
            if dt_from:
                filters.append(FlowLog.created_at >= dt_from)
        if date_to:
            dt_to = _parse_dt(date_to)
            if dt_to:
                # Include entire day when only a date is provided
                if date_to and len(date_to) == 10:
                    dt_to_upper = dt_to + timedelta(days=1)
                    filters.append(FlowLog.created_at < dt_to_upper)
                else:
                    filters.append(FlowLog.created_at <= dt_to)

        query = db.query(FlowLog).filter(and_(*filters)) if filters else db.query(FlowLog)
        if q:
            like = f"%{q}%"
            query = query.filter(FlowLog.description.ilike(like))

        total = query.count()
        rows = (
            query.order_by(FlowLog.created_at.desc())
            .offset((page - 1) * limit)
            .limit(limit)
            .all()
        )

        # Fetch customer names for wa_ids with missing names
        wa_ids_needing_names = set(r.wa_id for r in rows if r.wa_id and (not r.name or r.name.strip() == ''))
        customer_names = {}
        if wa_ids_needing_names:
            customers = db.query(Customer.wa_id, Customer.name).filter(Customer.wa_id.in_(wa_ids_needing_names)).all()
            customer_names = {c.wa_id: c.name for c in customers if c.name}

        def _row_to_dict(x: FlowLog) -> Dict[str, Any]:
            # Use flow_log name if available, otherwise fallback to customer name
            name = x.name
            if not name or name.strip() == '':
                name = customer_names.get(x.wa_id, '')
            return {
                "id": str(x.id),
                "created_at": x.created_at.isoformat() if x.created_at else None,
                "flow_type": x.flow_type,
                "name": name,
                "step": x.step,
                "status_code": x.status_code,
                "wa_id": x.wa_id,
                "description": x.description,
                "response_json": x.response_json,
            }

        return {
            "success": True,
            "total": total,
            "page": page,
            "limit": limit,
            "data": [_row_to_dict(r) for r in rows],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-pushed-leads")
def export_pushed_leads_excel(
    db: Session = Depends(get_db),
    flow_type: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
):
    """
    Export leads pushed to Zoho (stored in leads table) as an Excel workbook.
    Results can be filtered by flow type and date range.
    """
    try:
        dt_from = _parse_dt(date_from)
        dt_to = _parse_dt(date_to)
        dt_to_upper: Optional[datetime] = None
        if dt_to:
            # include entire day when only a date is provided
            if date_to and len(date_to) == 10:
                dt_to_upper = dt_to + timedelta(days=1)
            else:
                dt_to_upper = dt_to

        lead_query = db.query(Lead)
        if dt_from:
            lead_query = lead_query.filter(Lead.created_at >= dt_from)
        if dt_to_upper:
            lead_query = lead_query.filter(Lead.created_at < dt_to_upper)

        leads: List[Lead] = (
            lead_query.order_by(Lead.created_at.desc()).limit(20000).all()
        )

        if not leads:
            raise HTTPException(status_code=404, detail="No leads found for the selected filters.")

        log_filters = [
            FlowLog.step == "result",
            FlowLog.status_code == 200,
            FlowLog.response_json.isnot(None),
        ]
        if flow_type:
            log_filters.append(FlowLog.flow_type == flow_type)
        if dt_from:
            log_filters.append(FlowLog.created_at >= dt_from)
        if dt_to_upper:
            log_filters.append(FlowLog.created_at < dt_to_upper)

        flow_log_map: Dict[str, str] = {}
        wa_id_map: Dict[str, str] = {}

        log_rows = db.query(FlowLog).filter(and_(*log_filters)).all()
        for log in log_rows:
            if not log.response_json:
                continue
            lead_id: Optional[str] = None
            try:
                payload = json.loads(log.response_json)
                if isinstance(payload, dict):
                    if payload.get("success") and not payload.get("duplicate") and not payload.get("skipped"):
                        lead_id = payload.get("lead_id")
                        if not lead_id:
                            response_block = payload.get("response") or {}
                            data_list = response_block.get("data") or []
                            if data_list:
                                details = (data_list[0] or {}).get("details") or {}
                                lead_id = details.get("id")
            except Exception:
                continue

            if not lead_id:
                continue
            lead_id_str = str(lead_id)
            flow_log_map.setdefault(lead_id_str, log.flow_type or "")
            if log.wa_id:
                wa_id_map.setdefault(lead_id_str, log.wa_id)

        flow_type_labels = {
            "treatment": "Marketing",
            "lead_appointment": "Meta Ad Campaign",
        }

        def infer_flow_from_source(lead_source: Optional[str]) -> Optional[str]:
            if not lead_source:
                return None
            source_lower = lead_source.lower()
            if "business" in source_lower:
                return "treatment"
            if "facebook" in source_lower or "meta" in source_lower:
                return "lead_appointment"
            return None

        # Get customer info for pushed leads FIRST
        lead_wa_ids = [lead.wa_id for lead in leads if lead.wa_id]
        customer_map_pushed = {}
        if lead_wa_ids:
            customers = db.query(Customer).filter(Customer.wa_id.in_(lead_wa_ids)).all()
            customer_map_pushed = {c.wa_id: c for c in customers}

        # Get peer numbers for all leads
        peer_number_map = _get_peer_numbers_for_customers(db, lead_wa_ids)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Pushed Leads"
        worksheet.append(["Name", "Lead ID", "Lead Source", "Sub Source", "Phone Number", "Peer Number", "Type of Flow", "Created At"])

        rows_written = 0
        for lead in leads:
            lead_id = lead.zoho_lead_id
            if not lead_id:
                continue

            flow_code = flow_log_map.get(lead_id) or infer_flow_from_source(lead.lead_source)
            if flow_type:
                if not flow_code:
                    continue
                if flow_code != flow_type:
                    continue

            flow_label = flow_type_labels.get(flow_code or "", flow_code or "Unknown")

            # Prioritize Customer.name, then Lead names, then "Unknown"
            customer = customer_map_pushed.get(lead.wa_id) if lead.wa_id else None
            if customer and customer.name and customer.name.strip():
                name = customer.name.strip()
            else:
                name_parts = [lead.first_name or "", lead.last_name or ""]
                name = " ".join(part.strip() for part in name_parts if part and part.strip()).strip() or "Unknown"

            phone = lead.phone or lead.mobile or wa_id_map.get(lead_id) or ""
            peer_number = peer_number_map.get(lead.wa_id, "")

            worksheet.append([
                name,
                lead_id,
                lead.lead_source or "",
                lead.sub_source or "",
                phone,
                peer_number,
                flow_label,
                lead.created_at.isoformat() if lead.created_at else "",
            ])
            rows_written += 1

        if rows_written == 0:
            raise HTTPException(status_code=404, detail="No leads matched the selected filters.")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        suffix_parts = []
        if flow_type:
            suffix_parts.append(flow_type)
        if date_from:
            suffix_parts.append(f"from_{date_from}")
        if date_to:
            suffix_parts.append(f"to_{date_to}")
        suffix = ("_" + "_".join(suffix_parts)) if suffix_parts else ""

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=pushed_leads{suffix}.xlsx"
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-non-pushed-leads")
def export_non_pushed_leads_excel(
    db: Session = Depends(get_db),
    flow_type: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
):
    """
    Export leads that started a flow but were NOT pushed to Zoho (no Lead entry).
    Results can be filtered by flow type and date range.
    """
    try:
        dt_from = _parse_dt(date_from)
        dt_to = _parse_dt(date_to)
        dt_to_upper: Optional[datetime] = None
        if dt_to:
            if date_to and len(date_to) == 10:
                dt_to_upper = dt_to + timedelta(days=1)
            else:
                dt_to_upper = dt_to

        # Get all FlowLog entries that represent flow starts
        log_filters = [
            FlowLog.step.isnot(None),
            FlowLog.wa_id.isnot(None),
        ]
        if flow_type:
            log_filters.append(FlowLog.flow_type == flow_type)
        if dt_from:
            log_filters.append(FlowLog.created_at >= dt_from)
        if dt_to_upper:
            log_filters.append(FlowLog.created_at < dt_to_upper)

        # Get all unique wa_ids from FlowLog that have entries
        flow_log_wa_ids_result = (
            db.query(FlowLog.wa_id)
            .filter(and_(*log_filters))
            .distinct()
            .all()
        )
        flow_log_wa_ids_set = {row[0] for row in flow_log_wa_ids_result if row[0]}

        # Get all wa_ids that have Lead entries (pushed to Zoho)
        pushed_wa_ids_result = (
            db.query(Lead.wa_id)
            .distinct()
            .all()
        )
        pushed_wa_ids_set = {row[0] for row in pushed_wa_ids_result if row[0]}

        # Find wa_ids that are in FlowLog but NOT in Lead (non-pushed leads)
        non_pushed_wa_ids_set = flow_log_wa_ids_set - pushed_wa_ids_set
        non_pushed_wa_ids = list(non_pushed_wa_ids_set)

        if not non_pushed_wa_ids:
            raise HTTPException(status_code=404, detail="No non-pushed leads found for the selected filters.")

        # Get latest FlowLog entry for each non-pushed wa_id
        latest_log_subq = (
            db.query(
                FlowLog.wa_id,
                func.max(FlowLog.created_at).label('max_created_at')
            )
            .filter(FlowLog.wa_id.in_(non_pushed_wa_ids))
            .filter(and_(*log_filters))
            .group_by(FlowLog.wa_id)
            .subquery()
        )

        # Get full FlowLog entries with customer info
        results = (
            db.query(
                FlowLog,
                Customer.name,
                Customer.phone_1,
                Customer.email
            )
            .join(
                latest_log_subq,
                and_(
                    FlowLog.wa_id == latest_log_subq.c.wa_id,
                    FlowLog.created_at == latest_log_subq.c.max_created_at
                )
            )
            .outerjoin(Customer, FlowLog.wa_id == Customer.wa_id)
            .all()
        )

        # Get peer numbers for non-pushed leads
        non_pushed_wa_ids = [log.wa_id for log, _, _, _ in results if log.wa_id]
        peer_number_map = _get_peer_numbers_for_customers(db, non_pushed_wa_ids)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Non-Pushed Leads"
        worksheet.append([
            "Name", "Phone Number", "Email", "WA ID", "Peer Number", "Flow Type", "Last Step", 
            "Description", "Status Code", "Created At"
        ])

        rows_written = 0
        for log, customer_name, phone_1, email in results:
            # Prioritize Customer.name from join, then FlowLog.name, then "Unknown"
            if customer_name and customer_name.strip():
                name = customer_name.strip()
            elif log.name and log.name.strip():
                name = log.name.strip()
            else:
                name = "Unknown"
            phone = phone_1 or log.wa_id or ""
            peer_number = peer_number_map.get(log.wa_id, "") if log.wa_id else ""
            
            worksheet.append([
                name,
                phone,
                email or "",
                log.wa_id or "",
                peer_number,
                log.flow_type or "",
                log.step or "",
                log.description or "",
                log.status_code or "",
                log.created_at.isoformat() if log.created_at else "",
            ])
            rows_written += 1

        if rows_written == 0:
            raise HTTPException(status_code=404, detail="No non-pushed leads matched the selected filters.")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        suffix_parts = []
        if flow_type:
            suffix_parts.append(flow_type)
        if date_from:
            suffix_parts.append(f"from_{date_from}")
        if date_to:
            suffix_parts.append(f"to_{date_to}")
        suffix = ("_" + "_".join(suffix_parts)) if suffix_parts else ""

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=non_pushed_leads{suffix}.xlsx"
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-all-leads")
def export_all_leads_excel(
    db: Session = Depends(get_db),
    flow_type: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
):
    """
    Export ALL customers (both pushed and non-pushed) who messaged in the date range.
    Includes customers with messages but no FlowLog entries to ensure no contacts are missing.
    Results can be filtered by flow type and date range.
    """
    try:
        # Validate and parse dates
        if not date_from or not date_to:
            raise HTTPException(status_code=400, detail="Both date_from and date_to are required (YYYY-MM-DD)")
        
        dt_from = _parse_dt(date_from)
        dt_to = _parse_dt(date_to)
        
        if not dt_from or not dt_to:
            raise HTTPException(status_code=400, detail="Invalid date format. Please use YYYY-MM-DD format.")
        
        # Ensure dt_to includes the entire day
        if date_to and len(date_to) == 10:
            dt_to_upper = dt_to + timedelta(days=1)
        else:
            dt_to_upper = dt_to

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "All Customers"

        flow_type_labels = {
            "treatment": "Marketing",
            "lead_appointment": "Meta Ad Campaign",
        }

        def infer_flow_from_source(lead_source: Optional[str]) -> Optional[str]:
            if not lead_source:
                return None
            source_lower = lead_source.lower()
            if "business" in source_lower:
                return "treatment"
            if "facebook" in source_lower or "meta" in source_lower:
                return "lead_appointment"
            return None

        # Unified header with all columns
        worksheet.append([
            "Status", "Name", "Phone Number", "Email", "WA ID", "Lead ID", 
            "Lead Source", "Sub Source", "Peer Number", "Flow Type", 
            "Last Step", "Description", "Status Code", "Keywords", "Who Ended Chat First", "Message Count", "First Message", "Last Message", "Created At"
        ])
        
        # FIRST: Get ALL customers who messaged in the date range (to ensure none are missing)
        # Build message filters properly - only add date filters if dates are provided
        message_filters = [Message.sender_type == "customer"]
        
        if not dt_from or not dt_to_upper:
            raise HTTPException(status_code=400, detail="Both date_from and date_to are required (YYYY-MM-DD)")
        
        # Add date filters
        message_filters.append(Message.timestamp >= dt_from)
        message_filters.append(Message.timestamp < dt_to_upper)
        
        customers_with_messages_query = (
            db.query(Customer.wa_id)
            .join(Message, Customer.id == Message.customer_id)
            .filter(and_(*message_filters))
        )
        
        all_customers_wa_ids_result = customers_with_messages_query.distinct().all()
        all_customers_wa_ids_set = {row[0] for row in all_customers_wa_ids_result if row[0]}
        
        if not all_customers_wa_ids_set:
            raise HTTPException(status_code=404, detail="No customers found for the selected date range.")
        
        # Get message stats for all customers
        message_stats_query = (
            db.query(
                Customer.wa_id,
                func.count(Message.id).label("message_count"),
                func.min(Message.timestamp).label("first_message"),
                func.max(Message.timestamp).label("last_message")
            )
            .join(Message, Customer.id == Message.customer_id)
            .filter(Customer.wa_id.in_(list(all_customers_wa_ids_set)))
        )
        if dt_from:
            message_stats_query = message_stats_query.filter(Message.timestamp >= dt_from)
        if dt_to_upper:
            message_stats_query = message_stats_query.filter(Message.timestamp < dt_to_upper)
        
        message_stats_query = message_stats_query.group_by(Customer.wa_id)
        message_stats = {row[0]: {
            'count': row[1],
            'first': row[2],
            'last': row[3]
        } for row in message_stats_query.all()}

        lead_query = db.query(Lead)
        if dt_from:
            lead_query = lead_query.filter(Lead.created_at >= dt_from)
        if dt_to_upper:
            lead_query = lead_query.filter(Lead.created_at < dt_to_upper)

        leads: List[Lead] = (
            lead_query.order_by(Lead.created_at.desc()).limit(20000).all()
        )

        # Get all customer details
        all_customers = db.query(Customer).filter(Customer.wa_id.in_(list(all_customers_wa_ids_set))).all()
        customer_map_all = {c.wa_id: c for c in all_customers}
        
        # Get peer numbers for ALL customers
        all_wa_ids_list = list(all_customers_wa_ids_set)
        peer_number_map_all = _get_peer_numbers_for_customers(db, all_wa_ids_list)
        
        # Get keywords and who ended for ALL customers
        keywords_map_all = _extract_keywords_from_messages(db, all_wa_ids_list)
        who_ended_map_all = _get_who_ended_chat_first(db, all_wa_ids_list)
        
        # Get FlowLog info for customers who entered flows in date range
        flow_log_filters = [
            FlowLog.wa_id.in_(all_wa_ids_list),
        ]
        if dt_from:
            flow_log_filters.append(FlowLog.created_at >= dt_from)
        if dt_to_upper:
            flow_log_filters.append(FlowLog.created_at < dt_to_upper)
        if flow_type:
            flow_log_filters.append(FlowLog.flow_type == flow_type)
        
        flow_logs_all = db.query(FlowLog).filter(and_(*flow_log_filters)).all()
        
        # Get latest flow log for each customer
        customer_flow_map = {}
        for log in flow_logs_all:
            if not log.wa_id:
                continue
            if log.wa_id not in customer_flow_map:
                customer_flow_map[log.wa_id] = log
            else:
                # Keep the latest flow log
                if log.created_at and customer_flow_map[log.wa_id].created_at:
                    if log.created_at > customer_flow_map[log.wa_id].created_at:
                        customer_flow_map[log.wa_id] = log
        
        # Get pushed leads (customers who have Lead entries)
        lead_map_by_wa_id = {lead.wa_id: lead for lead in leads if lead.wa_id}
        pushed_wa_ids_set = set(lead_map_by_wa_id.keys())
        
        # Process ALL customers who messaged in date range
        rows_written = 0
        
        for wa_id in all_customers_wa_ids_set:
            customer = customer_map_all.get(wa_id)
            if not customer:
                continue
            
            # Get customer name (prioritize Customer.name)
            name = customer.name if customer.name and customer.name.strip() else "Unknown"
            phone = customer.phone_1 or wa_id or ""
            email = customer.email or ""
            peer_number = peer_number_map_all.get(wa_id, "") if wa_id else ""
            
            # Get message stats
            stats = message_stats.get(wa_id, {'count': 0, 'first': None, 'last': None})
            message_count = stats['count']
            first_message = stats['first']
            last_message = stats['last']
            
            # Get keywords and who ended
            keywords = keywords_map_all.get(wa_id, "") if wa_id else ""
            who_ended = who_ended_map_all.get(wa_id, "") if wa_id else ""
            
            # Determine if pushed or non-pushed
            is_pushed = wa_id in pushed_wa_ids_set
            status = "Pushed" if is_pushed else "Non-Pushed"
            
            # Get Lead info if pushed
            lead_id = ""
            lead_source = ""
            sub_source = ""
            lead_created_at = ""
            if is_pushed:
                lead = lead_map_by_wa_id.get(wa_id)
                if lead:
                    lead_id = lead.zoho_lead_id or ""
                    lead_source = lead.lead_source or ""
                    sub_source = lead.sub_source or ""
                    lead_created_at = lead.created_at.isoformat() if lead.created_at else ""
            
            # Get FlowLog info if available
            flow_log = customer_flow_map.get(wa_id)
            flow_type_str = ""
            last_step = ""
            description = ""
            status_code = ""
            
            if flow_log:
                flow_type_str = flow_type_labels.get(flow_log.flow_type or "", flow_log.flow_type or "")
                last_step = flow_log.step or ""
                description = flow_log.description or ""
                status_code = str(flow_log.status_code) if flow_log.status_code else ""
            
            # If flow_type filter is set, only include customers with matching flow type
            # Include customers who messaged even if they don't have a flow log (they might have messaged but not entered a flow)
            if flow_type:
                # If customer has a flow log, check if it matches the filter
                if flow_log:
                    if flow_log.flow_type != flow_type:
                        continue  # Skip if flow type doesn't match
                # If customer has no flow log but flow_type filter is set, 
                # we skip them as they didn't enter any flow (and we're filtering by flow type)
                else:
                    continue
            
            worksheet.append([
                status,
                name,
                phone,
                email,
                wa_id,
                lead_id,
                lead_source,
                sub_source,
                peer_number,
                flow_type_str,
                last_step,
                description,
                status_code,
                keywords,
                who_ended,
                message_count,
                first_message.isoformat() if first_message else "",
                last_message.isoformat() if last_message else "",
                lead_created_at if lead_created_at else (flow_log.created_at.isoformat() if flow_log and flow_log.created_at else "")
            ])
            rows_written += 1

        if rows_written == 0:
            raise HTTPException(status_code=404, detail="No customers found for the selected filters.")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        suffix_parts = []
        if flow_type:
            suffix_parts.append(flow_type)
        if date_from:
            suffix_parts.append(f"from_{date_from}")
        if date_to:
            suffix_parts.append(f"to_{date_to}")
        suffix = ("_" + "_".join(suffix_parts)) if suffix_parts else ""

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=all_leads{suffix}.xlsx"
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-all-customers-by-date")
def export_all_customers_by_date_excel(
    db: Session = Depends(get_db),
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    flow_type: Optional[str] = Query(None, description="Filter by flow type"),
):
    """
    Export ALL customers (both pushed and non-pushed) who messaged in the date range.
    Includes all customer data with peer numbers in a single Excel sheet.
    Ensures no customers are missing - includes all who messaged in the date range.
    """
    try:
        dt_from = _parse_dt(date_from)
        dt_to = _parse_dt(date_to)
        
        if not dt_from or not dt_to:
            raise HTTPException(status_code=400, detail="Both date_from and date_to are required (YYYY-MM-DD)")
        
        # Ensure dt_to includes the entire day
        if date_to and len(date_to) == 10:
            dt_to_upper = dt_to + timedelta(days=1)
        else:
            dt_to_upper = dt_to
        
        # Get ALL customers who messaged in the date range
        message_filters = [
            Message.timestamp >= dt_from,
            Message.timestamp < dt_to_upper,
            Message.sender_type == "customer"  # Only customer messages
        ]
        
        # Get all unique customers who messaged in date range
        customers_with_messages_query = (
            db.query(Customer.wa_id)
            .join(Message, Customer.id == Message.customer_id)
            .filter(and_(*message_filters))
            .distinct()
        )
        customers_with_messages_result = customers_with_messages_query.all()
        all_wa_ids = [row[0] for row in customers_with_messages_result if row[0]]
        
        if not all_wa_ids:
            raise HTTPException(status_code=404, detail="No customers found for the selected date range.")
        
        # Get all customer details
        customers = db.query(Customer).filter(Customer.wa_id.in_(all_wa_ids)).all()
        customer_map = {c.wa_id: c for c in customers}
        
        # Get message counts and timestamps for each customer in date range
        message_stats_query = (
            db.query(
                Customer.wa_id,
                func.count(Message.id).label("message_count"),
                func.min(Message.timestamp).label("first_message"),
                func.max(Message.timestamp).label("last_message")
            )
            .join(Message, Customer.id == Message.customer_id)
            .filter(Customer.wa_id.in_(all_wa_ids))
            .filter(and_(*message_filters))
            .group_by(Customer.wa_id)
        )
        message_stats = {row[0]: {
            'count': row[1],
            'first': row[2],
            'last': row[3]
        } for row in message_stats_query.all()}
        
        # Get FlowLog info for customers (if they entered a flow in date range)
        flow_log_filters = [
            FlowLog.wa_id.in_(all_wa_ids),
            FlowLog.created_at >= dt_from,
            FlowLog.created_at < dt_to_upper
        ]
        if flow_type:
            flow_log_filters.append(FlowLog.flow_type == flow_type)
        
        flow_logs_query = db.query(FlowLog).filter(and_(*flow_log_filters))
        flow_logs = flow_logs_query.all()
        
        # Group flow logs by wa_id - get latest flow info for each customer
        flow_log_map = {}
        for log in flow_logs:
            if not log.wa_id:
                continue
            if log.wa_id not in flow_log_map:
                flow_log_map[log.wa_id] = []
            flow_log_map[log.wa_id].append(log)
        
        # Get latest flow log for each customer
        customer_flow_info = {}
        for wa_id, logs in flow_log_map.items():
            latest_log = max(logs, key=lambda x: x.created_at if x.created_at else datetime.min)
            customer_flow_info[wa_id] = {
                'flow_type': latest_log.flow_type or "",
                'step': latest_log.step or "",
                'status_code': str(latest_log.status_code) if latest_log.status_code else "",
                'description': latest_log.description or ""
            }
        
        # Get peer numbers for all customers
        peer_number_map = _get_peer_numbers_for_customers(db, all_wa_ids)
        
        # Check which customers are pushed (have Lead entries)
        pushed_wa_ids_result = (
            db.query(Lead.wa_id)
            .filter(Lead.wa_id.in_(all_wa_ids))
            .distinct()
            .all()
        )
        pushed_wa_ids_set = {row[0] for row in pushed_wa_ids_result if row[0]}
        
        # Get Lead info for pushed customers
        leads_query = db.query(Lead).filter(Lead.wa_id.in_(list(pushed_wa_ids_set)))
        if dt_from:
            leads_query = leads_query.filter(Lead.created_at >= dt_from)
        if dt_to_upper:
            leads_query = leads_query.filter(Lead.created_at < dt_to_upper)
        leads = leads_query.all()
        lead_map = {lead.wa_id: lead for lead in leads}
        
        # Create workbook with single sheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "All Customers"
        
        # Header row
        worksheet.append([
            "Status", "Name", "Phone Number", "Email", "WA ID", "Peer Number",
            "Message Count", "First Message Time", "Last Message Time",
            "Flow Type", "Last Flow Step", "Flow Status", "Flow Description",
            "Lead ID", "Lead Source", "Sub Source", "Organization ID"
        ])
        
        rows_written = 0
        for wa_id in all_wa_ids:
            customer = customer_map.get(wa_id)
            if not customer:
                continue
            
            # Get customer name (prioritize Customer.name)
            name = customer.name if customer.name and customer.name.strip() else "Unknown"
            phone = customer.phone_1 or wa_id or ""
            email = customer.email or ""
            peer_number = peer_number_map.get(wa_id, "") if wa_id else ""
            
            # Get message stats
            stats = message_stats.get(wa_id, {'count': 0, 'first': None, 'last': None})
            message_count = stats['count']
            first_message = stats['first']
            last_message = stats['last']
            
            # Get flow info
            flow_info = customer_flow_info.get(wa_id, {
                'flow_type': "",
                'step': "",
                'status_code': "",
                'description': ""
            })
            
            # Determine status (Pushed or Non-Pushed)
            is_pushed = wa_id in pushed_wa_ids_set
            status = "Pushed" if is_pushed else "Non-Pushed"
            
            # Get Lead info if pushed
            lead_id = ""
            lead_source = ""
            sub_source = ""
            if is_pushed:
                lead = lead_map.get(wa_id)
                if lead:
                    lead_id = lead.zoho_lead_id or ""
                    lead_source = lead.lead_source or ""
                    sub_source = lead.sub_source or ""
            
            worksheet.append([
                status,
                name,
                phone,
                email,
                wa_id,
                peer_number,
                message_count,
                first_message.isoformat() if first_message else "",
                last_message.isoformat() if last_message else "",
                flow_info['flow_type'],
                flow_info['step'],
                flow_info['status_code'],
                flow_info['description'],
                lead_id,
                lead_source,
                sub_source,
                str(customer.organization_id) if customer.organization_id else ""
            ])
            rows_written += 1
        
        if rows_written == 0:
            raise HTTPException(status_code=404, detail="No data to export.")
        
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        suffix = f"_{date_from}_to_{date_to}" if date_from and date_to else ""
        if flow_type:
            suffix += f"_{flow_type}"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=all_customers_by_date{suffix}.xlsx"
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export customers: {str(e)}")


@router.get("/export-chats-without-push")
def export_chats_without_push_excel(
    db: Session = Depends(get_db),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
):
    """
    Export customers who have messages (chats) but no FlowLog entries (never entered a flow).
    Results can be filtered by date range.
    """
    try:
        dt_from = _parse_dt(date_from)
        dt_to = _parse_dt(date_to)
        dt_to_upper: Optional[datetime] = None
        if dt_to:
            if date_to and len(date_to) == 10:
                dt_to_upper = dt_to + timedelta(days=1)
            else:
                dt_to_upper = dt_to

        # Get all wa_ids that have messages (join Message with Customer)
        message_filters = []
        if dt_from:
            message_filters.append(Message.timestamp >= dt_from)
        if dt_to_upper:
            message_filters.append(Message.timestamp < dt_to_upper)

        customers_with_messages_query = (
            db.query(Customer.wa_id)
            .join(Message, Customer.id == Message.customer_id)
        )
        if message_filters:
            customers_with_messages_query = customers_with_messages_query.filter(and_(*message_filters))
        customers_with_messages_result = customers_with_messages_query.distinct().all()
        customers_with_messages_set = {row[0] for row in customers_with_messages_result if row[0]}

        # Get all wa_ids that have FlowLog entries (entered a flow)
        customers_with_flows_wa_ids = (
            db.query(FlowLog.wa_id)
            .distinct()
            .all()
        )
        customers_with_flows_set = {row[0] for row in customers_with_flows_wa_ids if row[0]}

        # Find customers with messages but NO FlowLog entries
        chats_without_flow_wa_ids = customers_with_messages_set - customers_with_flows_set

        if not chats_without_flow_wa_ids:
            raise HTTPException(status_code=404, detail="No chats without flow push found for the selected filters.")

        # Get customers
        customers_query = (
            db.query(Customer)
            .filter(Customer.wa_id.in_(list(chats_without_flow_wa_ids)))
        )

        if dt_from or dt_to_upper:
            # Also filter by customer creation date if date filters are provided
            customer_date_filters = []
            if dt_from:
                customer_date_filters.append(Customer.created_at >= dt_from)
            if dt_to_upper:
                customer_date_filters.append(Customer.created_at < dt_to_upper)
            if customer_date_filters:
                customers_query = customers_query.filter(and_(*customer_date_filters))

        customers = customers_query.order_by(Customer.created_at.desc()).limit(20000).all()

        if not customers:
            raise HTTPException(status_code=404, detail="No chats without flow push found for the selected filters.")

        # Get latest message for each customer (join via customer_id)
        customer_wa_ids = [c.wa_id for c in customers if c.wa_id]
        if not customer_wa_ids:
            message_map = {}
        else:
            latest_message_subq = (
                db.query(
                    Message.customer_id,
                    func.max(Message.timestamp).label('max_timestamp')
                )
                .join(Customer, Message.customer_id == Customer.id)
                .filter(Customer.wa_id.in_(customer_wa_ids))
                .group_by(Message.customer_id)
                .subquery()
            )

            latest_messages = (
                db.query(
                    Customer.wa_id,
                    Message.body,
                    Message.timestamp
                )
                .select_from(Message)
                .join(
                    latest_message_subq,
                    and_(
                        Message.customer_id == latest_message_subq.c.customer_id,
                        Message.timestamp == latest_message_subq.c.max_timestamp
                    )
                )
                .join(Customer, Message.customer_id == Customer.id)
                .all()
            )
            message_map = {wa_id: (body, timestamp) for wa_id, body, timestamp in latest_messages}

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Chats Without Push"
        worksheet.append([
            "Name", "Phone Number", "Email", "WA ID", "Last Message", "Last Message Time", "Created At"
        ])

        rows_written = 0
        for customer in customers:
            last_msg_body, last_msg_time = message_map.get(customer.wa_id, ("", None))
            
            name = customer.name or "Unknown"
            phone = customer.phone_1 or customer.wa_id or ""
            
            worksheet.append([
                name,
                phone,
                customer.email or "",
                customer.wa_id or "",
                (last_msg_body[:100] if last_msg_body else ""),  # Truncate long messages
                last_msg_time.isoformat() if last_msg_time else "",
                customer.created_at.isoformat() if customer.created_at else "",
            ])
            rows_written += 1

        if rows_written == 0:
            raise HTTPException(status_code=404, detail="No chats without flow push matched the selected filters.")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        suffix_parts = []
        if date_from:
            suffix_parts.append(f"from_{date_from}")
        if date_to:
            suffix_parts.append(f"to_{date_to}")
        suffix = ("_" + "_".join(suffix_parts)) if suffix_parts else ""

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=chats_without_push{suffix}.xlsx"
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/lead-counts")
def get_lead_counts(
    db: Session = Depends(get_db),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
):
    """
    Return counts of leads pushed to Zoho, grouped by flow type.
    """
    try:
        dt_from = _parse_dt(date_from)
        dt_to = _parse_dt(date_to)
        dt_to_upper: Optional[datetime] = None
        if dt_to:
            if date_to and len(date_to) == 10:
                dt_to_upper = dt_to + timedelta(days=1)
            else:
                dt_to_upper = dt_to

        lead_query = db.query(Lead)
        if dt_from:
            lead_query = lead_query.filter(Lead.created_at >= dt_from)
        if dt_to_upper:
            lead_query = lead_query.filter(Lead.created_at < dt_to_upper)

        leads: List[Lead] = lead_query.order_by(Lead.created_at.desc()).limit(20000).all()

        if not leads:
            return {
                "success": True,
                "overall": {"total_leads": 0, "unique_customers": 0},
                "treatment": {"total_leads": 0, "unique_customers": 0},
                "lead_appointment": {"total_leads": 0, "unique_customers": 0},
            }

        log_filters = [
            FlowLog.step == "result",
            FlowLog.status_code == 200,
            FlowLog.response_json.isnot(None),
        ]
        if dt_from:
            log_filters.append(FlowLog.created_at >= dt_from)
        if dt_to_upper:
            log_filters.append(FlowLog.created_at < dt_to_upper)

        flow_log_map: Dict[str, str] = {}
        log_rows = db.query(FlowLog).filter(and_(*log_filters)).all()
        for log in log_rows:
            if not log.response_json:
                continue
            try:
                payload = json.loads(log.response_json)
                if not isinstance(payload, dict):
                    continue
                if not payload.get("success"):
                    continue
                if payload.get("duplicate") or payload.get("skipped"):
                    continue
                lead_id = payload.get("lead_id")
                if not lead_id:
                    response_block = payload.get("response") or {}
                    data_list = response_block.get("data") or []
                    if data_list:
                        details = (data_list[0] or {}).get("details") or {}
                        lead_id = details.get("id")
                if not lead_id:
                    continue
                flow_log_map[str(lead_id)] = log.flow_type or ""
            except Exception:
                continue

        def infer_flow_from_source(lead_source: Optional[str]) -> Optional[str]:
            if not lead_source:
                return None
            lower = lead_source.lower()
            if "business" in lower:
                return "treatment"
            if "facebook" in lower or "meta" in lower:
                return "lead_appointment"
            return None

        counters = {
            "overall": {"total_leads": 0, "customers": set()},
            "treatment": {"total_leads": 0, "customers": set()},
            "lead_appointment": {"total_leads": 0, "customers": set()},
        }  # type: Dict[str, Dict[str, Any]]

        for lead in leads:
            lead_id = lead.zoho_lead_id
            if not lead_id:
                continue

            flow_code = flow_log_map.get(lead_id) or infer_flow_from_source(lead.lead_source)

            counters["overall"]["total_leads"] += 1
            if lead.wa_id:
                counters["overall"]["customers"].add(lead.wa_id)

            if flow_code in {"treatment", "lead_appointment"}:
                counters[flow_code]["total_leads"] += 1
                if lead.wa_id:
                    counters[flow_code]["customers"].add(lead.wa_id)

        def pack(flow_key: str) -> Dict[str, int]:
            entry = counters.get(flow_key) or {"total_leads": 0, "customers": set()}
            customers: Set[str] = entry.get("customers", set())  # type: ignore
            return {
                "total_leads": int(entry.get("total_leads", 0)),
                "unique_customers": len(customers),
            }

        return {
            "success": True,
            "overall": pack("overall"),
            "treatment": pack("treatment"),
            "lead_appointment": pack("lead_appointment"),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/lead-statistics")
def get_lead_statistics(
    db: Session = Depends(get_db),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
):
    """
    Return detailed lead statistics including:
    - Total leads pushed to Zoho
    - Duplicates
    - Errors
    - Followups
    - Completed
    - Leads generated (followups + completed)
    """
    try:
        dt_from = _parse_dt(date_from)
        dt_to = _parse_dt(date_to)
        dt_to_upper: Optional[datetime] = None
        if dt_to:
            if date_to and len(date_to) == 10:
                dt_to_upper = dt_to + timedelta(days=1)
            else:
                dt_to_upper = dt_to

        # Get all FlowLog entries in date range
        log_filters = []
        if dt_from:
            log_filters.append(FlowLog.created_at >= dt_from)
        if dt_to_upper:
            log_filters.append(FlowLog.created_at < dt_to_upper)

        query = db.query(FlowLog).filter(and_(*log_filters)) if log_filters else db.query(FlowLog)
        all_logs: List[FlowLog] = query.order_by(FlowLog.wa_id.asc(), FlowLog.created_at.asc()).all()

        # Group logs by wa_id
        groups: Dict[str, List[FlowLog]] = {}
        for log in all_logs:
            if not log.wa_id:
                continue
            groups.setdefault(log.wa_id, []).append(log)

        # Keywords for classification
        duplicate_keywords = ["duplicate", "duplicate avoided", "duplicate detected", "skipped", "already exists"]
        error_keywords = ["error", "failed", "exception", "timeout"]
        followup_keywords = ["follow-up", "follow up", "followup"]
        completion_keywords = ["thank you", "thankyou", "thank-you", "completed", "lead created"]

        # Counters
        total_leads_pushed = 0  # Actual leads in Lead table
        duplicates = 0
        errors = 0
        followups = 0
        completed = 0

        # Count actual leads pushed to Zoho (from Lead table)
        lead_query = db.query(Lead)
        if dt_from:
            lead_query = lead_query.filter(Lead.created_at >= dt_from)
        if dt_to_upper:
            lead_query = lead_query.filter(Lead.created_at < dt_to_upper)
        total_leads_pushed = lead_query.count()

        # Analyze each group
        for wa_id, logs in groups.items():
            sorted_logs = sorted(logs, key=lambda x: x.created_at or datetime.min, reverse=True)
            
            # Check for duplicates
            has_duplicate = False
            for log in sorted_logs:
                desc = (log.description or "").lower()
                response_json_str = log.response_json or ""
                
                # Check description for duplicate keywords
                if any(keyword in desc for keyword in duplicate_keywords):
                    has_duplicate = True
                    break
                
                # Check response_json for duplicate flag
                try:
                    if response_json_str:
                        payload = json.loads(response_json_str)
                        if isinstance(payload, dict) and (payload.get("duplicate") or payload.get("skipped")):
                            has_duplicate = True
                            break
                except Exception:
                    pass
            
            if has_duplicate:
                duplicates += 1
                continue

            # Check for errors
            has_error = False
            for log in sorted_logs:
                if log.status_code and log.status_code >= 400:
                    has_error = True
                    break
                desc = (log.description or "").lower()
                if any(keyword in desc for keyword in error_keywords):
                    has_error = True
                    break
            
            if has_error:
                errors += 1
                continue

            # Check for completion
            has_completion = False
            for log in sorted_logs:
                desc = (log.description or "").lower()
                step = (log.step or "").lower()
                status = log.status_code or 0
                
                # Check if step is "result" with 200 status
                if status >= 200 and status < 300 and step == "result":
                    # Verify it's not a duplicate by checking response_json
                    response_json_str = log.response_json or ""
                    try:
                        if response_json_str:
                            payload = json.loads(response_json_str)
                            if isinstance(payload, dict) and payload.get("success") and not payload.get("duplicate") and not payload.get("skipped"):
                                has_completion = True
                                break
                    except Exception:
                        pass
                
                # Check for completion keywords
                if any(keyword in desc for keyword in completion_keywords):
                    has_completion = True
                    break
            
            if has_completion:
                completed += 1
                continue

            # Check for followup
            has_followup = False
            for log in sorted_logs:
                step = (log.step or "").lower()
                desc = (log.description or "").lower()
                if any(keyword in step or keyword in desc for keyword in followup_keywords):
                    has_followup = True
                    break
            
            if has_followup:
                followups += 1
                continue

        # Leads generated = followups + completed
        leads_generated = followups + completed

        return {
            "success": True,
            "total_leads_pushed": total_leads_pushed,
            "duplicates": duplicates,
            "errors": errors,
            "followups": followups,
            "completed": completed,
            "leads_generated": leads_generated,  # followups + completed
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/completion-counts")
def get_completion_counts(
    db: Session = Depends(get_db),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
):
    """
    Get counts of customers who completed full flows (reached 'thank you' stage).
    A flow is considered completed if any log description contains a thank-you message.
    If the flow only has follow-up messages (and no thank-you), it is counted as non-completed.
    """
    try:
        filters = []
        if date_from:
            dt_from = _parse_dt(date_from)
            if dt_from:
                filters.append(FlowLog.created_at >= dt_from)
        if date_to:
            dt_to = _parse_dt(date_to)
            if dt_to:
                filters.append(FlowLog.created_at <= dt_to)

        query = db.query(FlowLog).filter(and_(*filters)) if filters else db.query(FlowLog)
        rows: List[FlowLog] = query.order_by(FlowLog.wa_id.asc(), FlowLog.created_at.asc()).all()

        treatment_groups: Dict[str, List[FlowLog]] = {}
        lead_appointment_groups: Dict[str, List[FlowLog]] = {}

        for r in rows:
            if not r.wa_id:
                continue
            if r.flow_type == "treatment":
                treatment_groups.setdefault(r.wa_id, []).append(r)
            elif r.flow_type == "lead_appointment":
                lead_appointment_groups.setdefault(r.wa_id, []).append(r)

        thank_you_keywords = [
            "thank you",
            "thankyou",
            "thank-you",
            "✅ thank you",
            "thank you!",
            "thank you 😊",
            "thank you 🙏",
        ]
        follow_up_keywords = [
            "follow-up",
            "follow up",
            "followup",
        ]

        def analyze_groups(groups: Dict[str, List[FlowLog]]) -> Dict[str, int]:
            totals = {
                "total": 0,
                "completed": 0,
                "followups": 0,
            }

            for wa_id, logs in groups.items():
                totals["total"] += 1
                descriptions = [
                    (log.description or "").lower()
                    for log in logs
                ]

                has_thank_you = any(
                    any(keyword in desc for keyword in thank_you_keywords)
                    for desc in descriptions
                )

                has_follow_up = any(
                    any(keyword in desc for keyword in follow_up_keywords)
                    for desc in descriptions
                )

                if has_thank_you:
                    totals["completed"] += 1
                elif has_follow_up:
                    totals["followups"] += 1
                # Remaining flows will be treated as pending (total - completed - followups)

            return totals

        treatment_totals = analyze_groups(treatment_groups)
        lead_totals = analyze_groups(lead_appointment_groups)

        def build_result(totals: Dict[str, int]) -> Dict[str, int]:
            pending = max(totals["total"] - totals["completed"] - totals["followups"], 0)
            return {
                "total": totals["total"],
                "completed": totals["completed"],
                "followups": totals["followups"],
                "pending": pending,
            }

        return {
            "success": True,
            "treatment": build_result(treatment_totals),
            "lead_appointment": build_result(lead_totals),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class BatchLastStepRequest(BaseModel):
    """Request model for batch last-step lookup"""
    wa_ids: List[str]
    flow_type: Optional[str] = "lead_appointment"


@router.post("/last-step/batch")
def get_last_step_batch(
    request: BatchLastStepRequest,
    db: Session = Depends(get_db),
):
    """
    Get the last step reached for multiple customers in a single request.
    Much faster than calling /last-step/{wa_id} for each customer individually.

    Request body:
    {
        "wa_ids": ["919740498236", "919876543210", ...],
        "flow_type": "lead_appointment"  // or "treatment"
    }

    Returns:
    {
        "success": true,
        "total": 100,
        "results": {
            "919740498236": { "last_step": "city_selection", ... },
            "919876543210": { "last_step": null, "message": "No data found" },
            ...
        }
    }
    """
    try:
        wa_ids = request.wa_ids
        flow_type = request.flow_type

        if not wa_ids:
            return {"success": True, "total": 0, "results": {}}

        # Limit to prevent abuse
        if len(wa_ids) > 1000:
            raise HTTPException(status_code=400, detail="Maximum 1000 wa_ids allowed per request")

        # Normalize flow_type
        normalized_flow_type = flow_type
        if flow_type == "treatment_flow":
            normalized_flow_type = "treatment"
        elif flow_type == "lead_appointment_flow":
            normalized_flow_type = "lead_appointment"

        valid_steps = ["entry", "city_selection", "treatment", "concern_list", "last_step"]

        # Single query to get the latest step for ALL wa_ids at once
        # Using a subquery to get the max created_at per wa_id, then joining back
        from sqlalchemy import func
        from sqlalchemy.orm import aliased

        # Subquery: get max created_at per wa_id for valid steps
        subq = (
            db.query(
                FlowLog.wa_id,
                func.max(FlowLog.created_at).label("max_created_at")
            )
            .filter(
                and_(
                    FlowLog.wa_id.in_(wa_ids),
                    FlowLog.flow_type == normalized_flow_type,
                    FlowLog.step.in_(valid_steps)
                )
            )
            .group_by(FlowLog.wa_id)
            .subquery()
        )

        # Main query: join back to get full log data
        logs = (
            db.query(FlowLog)
            .join(
                subq,
                and_(
                    FlowLog.wa_id == subq.c.wa_id,
                    FlowLog.created_at == subq.c.max_created_at
                )
            )
            .filter(
                and_(
                    FlowLog.flow_type == normalized_flow_type,
                    FlowLog.step.in_(valid_steps)
                )
            )
            .all()
        )

        # Build results map
        results = {}
        found_wa_ids = set()

        for log in logs:
            found_wa_ids.add(log.wa_id)
            results[log.wa_id] = {
                "last_step": log.step,
                "step_name": log.step,
                "reached_at": log.created_at.isoformat() if log.created_at else None,
                "customer_name": log.name,
                "description": log.description,
            }

        # Add null entries for wa_ids not found
        for wa_id in wa_ids:
            if wa_id not in found_wa_ids:
                results[wa_id] = {
                    "last_step": None,
                    "message": "No step data found for this customer"
                }

        return {
            "success": True,
            "total": len(wa_ids),
            "found": len(found_wa_ids),
            "flow_type": normalized_flow_type,
            "results": results
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/last-step/{wa_id}")
def get_last_step_reached(
    wa_id: str,
    db: Session = Depends(get_db),
    flow_type: Optional[str] = Query("lead_appointment", description="Flow type to check"),
):
    """
    Get the last step reached by a customer in a flow before follow-ups.
    Returns the most recent step from: entry, city_selection, treatment, concern_list, last_step
    
    Flow types:
    - "lead_appointment" or "lead_appointment_flow" -> searches for "lead_appointment"
    - "treatment" or "treatment_flow" -> searches for "treatment"
    """
    try:
        # Normalize flow_type (handle both "treatment_flow" and "treatment", etc.)
        normalized_flow_type = flow_type
        if flow_type == "treatment_flow":
            normalized_flow_type = "treatment"
        elif flow_type == "lead_appointment_flow":
            normalized_flow_type = "lead_appointment"
        
        # Valid steps for both flows
        valid_steps = ["entry", "city_selection", "treatment", "concern_list", "last_step"]
        
        # Query for the most recent valid step for this customer
        # Filter by step and description pattern (more flexible - matches any description containing "Last step reached")
        log = (
            db.query(FlowLog)
            .filter(
                and_(
                    FlowLog.wa_id == wa_id,
                    FlowLog.flow_type == normalized_flow_type,
                    FlowLog.step.in_(valid_steps),
                    or_(
                        FlowLog.description.like("%Last step reached:%"),
                        FlowLog.description == None  # Also include logs without description if step matches
                    )
                )
            )
            .order_by(FlowLog.created_at.desc())
            .first()
        )
        
        # If no log found with description filter, try without description filter (in case format changed)
        if not log:
            log = (
                db.query(FlowLog)
                .filter(
                    and_(
                        FlowLog.wa_id == wa_id,
                        FlowLog.flow_type == normalized_flow_type,
                        FlowLog.step.in_(valid_steps)
                    )
                )
                .order_by(FlowLog.created_at.desc())
                .first()
            )
        
        if not log:
            # Debug: Check if there are any logs at all for this customer
            any_logs = db.query(FlowLog).filter(
                and_(
                    FlowLog.wa_id == wa_id,
                    FlowLog.flow_type == normalized_flow_type
                )
            ).count()
            
            return {
                "success": True,
                "wa_id": wa_id,
                "flow_type": flow_type,
                "normalized_flow_type": normalized_flow_type,
                "last_step": None,
                "message": "No step data found for this customer",
                "debug": {
                    "total_logs_for_customer": any_logs,
                    "valid_steps_checked": valid_steps
                }
            }
        
        return {
            "success": True,
            "wa_id": wa_id,
            "flow_type": flow_type,
            "normalized_flow_type": normalized_flow_type,
            "last_step": log.step,
            "step_name": log.step,
            "reached_at": log.created_at.isoformat() if log.created_at else None,
            "customer_name": log.name,
            "description": log.description,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{log_id}")
def get_flow_log(
    log_id: UUID,
    db: Session = Depends(get_db),
):
    try:
        log = db.query(FlowLog).filter(FlowLog.id == log_id).first()
        if not log:
            raise HTTPException(status_code=404, detail="Log not found")

        return {
            "success": True,
            "data": {
                "id": str(log.id),
                "created_at": log.created_at.isoformat() if log.created_at else None,
                "flow_type": log.flow_type,
                "name": log.name,
                "step": log.step,
                "status_code": log.status_code,
                "wa_id": log.wa_id,
                "description": log.description,
                "response_json": log.response_json,
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))